import pandas as pd
import string
from openpyxl import load_workbook
from  openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

#detectar los datos a operar
def automatizar_excel(nombre_archivo):
    archivo_excel = pd.read_excel('supermarket_sales.xlsx')
    # print(archivo_excel[['Gender', 'Product line','Total']])
    tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    # print(tabla_pivote)
    mes_extension = nombre_archivo.split('_')[1]
    tabla_pivote.to_excel(f'sales_{mes_extension}', startrow=4, sheet_name='report')
    #Detectar las columnas a manejar
    wb = load_workbook(f'sales_{mes_extension}')
    pestaña = wb['report']
    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    # print(min_col)
    # print(max_col)
    # print(min_fila)
    # print(max_fila)

    #graficas
    barchart = BarChart()

    data = Reference(pestaña, min_col=min_col+1 , max_col=max_col , max_row=max_fila, min_row=min_fila)
    categorias = Reference(pestaña , min_col=min_col,max_col=min_col , max_row=max_fila , min_row=min_fila+1)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)


    pestaña.add_chart(barchart, 'B12')
    barchart.style = 2
    barchart.title = 'ventas'

    # pestaña['B8'] = '=SUM(B6:B7)'
    # pestaña['B8'].style = 'Currency'

    abecedario = list(string.ascii_uppercase)
    abecedario_execel = abecedario[0:max_col]

    for i in abecedario_execel:
        if i!='A':
            pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
            pestaña[f'{i}{max_fila+1}'].style = 'Currency'
            
    pestaña[f'{abecedario_execel[0]}{max_fila+1}'] = "Total"
    pestaña['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    pestaña['A2'] = mes

    pestaña['A1'].font = Font('Arial', bold=True, size=20)
    pestaña['A2'].font = Font('Arial', bold=True, size=12)


    wb.save(f'sales_{mes_extension}')
    return
automatizar_excel('sales_febrero.xlsx')